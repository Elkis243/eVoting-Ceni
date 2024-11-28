from openpyxl import load_workbook # type: ignore
from django.shortcuts import render, redirect, get_object_or_404
from django.views import View
from elector.models import *
from django.contrib import messages
from administration.models import *
from eVoting_Ceni.utils import hash_message
import logging
from django.db import transaction
from django.db.models import *
from django.db.models.functions import Coalesce
from django.utils import timezone
from eVoting_Ceni.utils import *
from django.conf import settings
from .models import *
import pytz
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required


logger = logging.getLogger("my_logger")

class ElectionView(View):
    @method_decorator(login_required)
    def get(self, request):
        elections = Election.objects.order_by("-end_date")
        election_exist = Election.objects.exists()
        number_elections = Election.objects.count()
        created_election_exist = Election.objects.filter(status='created').exists()
        active_election_exist = Election.objects.filter(status='active').exists()
        closed_election_exist = Election.objects.filter(status="closed").exists()
        number_elections_published = Election.objects.filter(status="published").count()
        no_existing_election = (number_elections == number_elections_published or number_elections == 0) and not(
            created_election_exist or active_election_exist or closed_election_exist
        )
        
        election = Election.objects.filter(status='created').first()
        if election is not None:
            count_list_electoral = ElectoralList.objects.filter(election=election).count()
        else:
            count_list_electoral = 0

        return render(request, 'administration/election.html',{
                "name": "election",
                "elections": elections,
                "election_exist": election_exist,
                "no_existing_election": no_existing_election,
                'count_list_electoral': count_list_electoral
            },
        )

    def post(self, request):
        start_date = datetime.fromisoformat(request.POST.get("start_date"))
        end_date = datetime.fromisoformat(request.POST.get("end_date"))
        
        if end_date < start_date :
            messages.error(request,"La date de fin que vous avez choisie est antérieure à la date de début !")
            return redirect("election")
        
        if start_date.year != end_date.year:
            messages.error(request, "Les dates de début et de fin doivent être dans la même année. Veuillez vérifier les années saisies !")
            return redirect("election")
        
        if start_date == end_date:
            messages.error(request, "La date de fin ne peut pas être identique à la date de début. Veuillez saisir des dates différentes !")
            return redirect("election")
        
        election_code = f"{settings.HASH_SALT}{end_date.year}"
        if Election.objects.filter(election_code=election_code).exists():
            messages.error(request, "Une élection pour cette année existe déjà. Veuillez vérifier les informations ou utiliser une autre année !")
            return redirect("election")
        
        local_tz = pytz.timezone('Africa/Kinshasa')
        election_name = f"Election {end_date.year}"
        local_start_date = local_tz.localize(start_date)
        local_end_date = local_tz.localize(end_date)
        
        election = Election.objects.create(
            election_code=election_code,
            election_name=election_name,
            start_date=local_start_date,
            end_date=local_end_date
        )
        logger.info(f"L'élection {election} a été créée avec succès.")
        messages.success(request, "L'élection a été créée avec succès. Vous pouvez maintenant poursuivre les autres étapes !")
        return redirect("election")
    
class EditElection(View):
    @method_decorator(login_required)
    def get(self, request, pk):
        try:
            election = get_object_or_404(Election, pk=pk)
            if election.start_date and election.end_date:
                election.start_date = election.start_date.strftime("%Y-%m-%dT%H:%M")
                election.end_date = election.end_date.strftime("%Y-%m-%dT%H:%M")
        except Election.DoesNotExist as e:
            logger.error(f"election does not exist : {e}")
            messages.error(request, "Une erreur inattendue s'est produite !")
            return redirect("election")
        
        return render(request, 'administration/edit_election.html', {
            'election': election
        })

    def post(self, request, pk):
        start_date = datetime.fromisoformat(request.POST.get("start_date"))
        end_date = datetime.fromisoformat(request.POST.get("end_date"))
        
        if end_date < start_date :
            messages.error(request,"La date de fin que vous avez choisie est antérieure à la date de début !")
            return redirect("election")
        
        if start_date.year != end_date.year:
            messages.error(request, "Les dates de début et de fin doivent être dans la même année. Veuillez vérifier les années saisies !")
            return redirect("election")
        
        if start_date == end_date:
            messages.error(request, "La date de fin ne peut pas être identique à la date de début. Veuillez saisir des dates différentes !")
            return redirect("election")
        
        local_tz = pytz.timezone('Africa/Kinshasa')
        election_name = f"Election {end_date.year}"
        local_start_date = local_tz.localize(start_date)
        local_end_date = local_tz.localize(end_date)

        try:
            election = get_object_or_404(Election, pk=pk)
            election.election_name = election_name
            election.start_date = local_start_date
            election.end_date = local_end_date
            election.save()
            logger.info(f"L'élection {election} a été modifiée avec succès.")
            messages.success(request, "Les modifications ont été enregistrées avec succès !")
            return redirect("election")
        except Election.DoesNotExist as e:
            logger.error(f"L'élection n'existe pas : {e}")
            messages.error(request, "L'élection demandée n'existe pas. Veuillez vérifier et réessayer !")
            return redirect("election")

@login_required
def cancelVoting(request, pk):
    try:
        election = get_object_or_404(Election, pk=pk)
        election.delete()
        logger.info(f"L'élection {election} a été annulée avec succès")
        messages.success(request, "L'élection a été annulée avec succès !")
        return redirect("election")
    except Election.DoesNotExist as e:
        logger.error(f"L'élection n'existe pas : {e}")
        messages.error(request, "Une erreur inattendue s'est produite !")
        return redirect("election")

@login_required
def import_list(request, pk):
    try:
        election = Election.objects.get(pk=pk, status='created')
    except Election.DoesNotExist as e:
        logger.error(f"Erreur lors de l'importation du fichier électoral : {e}")
        messages.error(request, "Une petite erreur est survenue lors de l'importation. Veuillez réessayer, s'il vous plaît !")
        return redirect('election')
    
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, "Le fichier importé n'est pas au format valide. Veuillez réessayer, s'il vous plaît !")
            return redirect('import_list', pk)
        
        try:
            wb = load_workbook(excel_file)
        except Exception as e:
            logger.error(f"Erreur lors de l'importation du fichier électoral : {e}")
            messages.error(request, "Une petite erreur est survenue lors de l'importation. Veuillez réessayer, s'il vous plaît !")
            return redirect('import_list', pk)
        
        ws = wb.active
        if ws.max_row < 2:
            messages.error(request, "Une petite erreur est survenue lors de l'importation. Le fichier importé ne contient aucune donnée !")
            return redirect('import_list', pk)
        
        expected_columns = ['national_number', 'email', 'country']
        file_columns = [cell.value for cell in ws[1]]
        
        if set(file_columns) != set(expected_columns):
            messages.error(request, "Les données fournies sont invalides. Merci de vérifier et de réessayer !")
            return redirect('import_list', pk)
        
        national_number_index = file_columns.index('national_number')
        email_index = file_columns.index('email')
        country_index = file_columns.index('country')
        
        national_numbers = []
        emails = []
        data_to_import = []
        
        for row in ws.iter_rows(min_row=2):
            national_number = row[national_number_index].value
            email = row[email_index].value
            country = row[country_index].value
            
            if not (isinstance(national_number, int) and isinstance(email, str) and isinstance(country, str)):
                messages.error(request, "Les données fournies sont invalides. Merci de vérifier et de réessayer !")
                return redirect('import_list', pk)
            
            national_number_str = str(national_number)
            if national_number_str in national_numbers or email in emails:
                messages.error(request, "Le fichier contient des doublons de numéros nationaux ou d'emails !")
                return redirect('import_list', pk)
            
            national_numbers.append(national_number_str)
            emails.append(email)
            data_to_import.append((national_number_str, email, country))
        
        existing_national_numbers = set(ElectoralList.objects.filter(national_number__in=national_numbers, election=election).values_list('national_number', flat=True))
        existing_emails = set(ElectoralList.objects.filter(email__in=emails, election=election).values_list('email', flat=True))
        
        if existing_national_numbers or existing_emails:
            messages.error(request, "Le fichier contient des données déjà présentes dans le système !")
            return redirect('import_list', pk)
        
        try:
            with transaction.atomic():
                for national_number, email, country in data_to_import:
                    ElectoralList.objects.create(
                        national_number=national_number,
                        email=email,
                        country=country,
                        election=election
                    )
            messages.success(request, "Les données ont été importées avec succès !")
            return redirect('election')
        except Exception as e:
            logger.error(f"Erreur lors de l'importation du fichier électoral : {e}")
            messages.error(request, "Une erreur est survenue lors de l'importation des données.")
            return redirect('import_list', pk)
    return render(request, "administration/import_list.html", {
            'election': election
        })

class CandidateView(View):
    @method_decorator(login_required)
    def get(self, request, pk):
        return render(request, 'administration/candidate.html', {
            'name': 'candidate'
        })

    def post(self, request, pk):
        name = request.POST.get("username")
        last_name = request.POST.get("last_name")
        first_name = request.POST.get("first_name")
        picture = request.FILES['picture']
        national_number = request.POST.get("national_number")
        national_number_hash = hash_message(national_number)
        party = request.POST.get("party")
        
        if Candidate.objects.filter(national_number=national_number_hash, election__pk=pk).exists():
            messages.error(request, "Ce candidat est déjà inscrit pour cette élection !")
            return redirect("election")
        
        try:
            election = get_object_or_404(Election, pk=pk)
            with transaction.atomic():
                candidate = Candidate.objects.create(
                    name=name,
                    last_name=last_name,
                    first_name=first_name,
                    picture=picture,
                    national_number=national_number,
                    party=party,
                    election=election
                )
                logger.info(f"Le candidat {candidate} a été créé avec succès.")
                messages.success(request, "Le candidat a été enregistré avec succès !")
                return redirect("election")
        except Exception as e:
            logger.error(f"L'élection n'existe pas : {e}")
            messages.error(request, "Une erreur inattendue s'est produite !")
            return redirect("election")

@login_required
def deleteCandidate(request, pk):
    try:
        candidate = get_object_or_404(Candidate, pk=pk)
        candidate.delete()
        logger.info(f"{candidate} deleted successfully.")
        messages.success(request, "Le candidat a été retiré avec succès !")
        return redirect("election")
    except Candidate.DoesNotExist as e:
        logger.error(f"candidate does not exist : {e}")
        messages.error(request, "Une erreur inattendue s'est produite !")
        return redirect("election")

@login_required
def activeVoting(request, pk):
    try:
        election = get_object_or_404(Election, pk=pk)
        election.status = "active"
        election.save()
        logger.info(f"L'élection {election} a été activée avec succès")
        messages.success(request, "L'élection a été activée avec succès !")
        return redirect("election")
    except Election.DoesNotExist as e:
        logger.error(f"L'élection n'existe pas : {e}")
        messages.error(request, "Une erreur inattendue s'est produite !")
        return redirect("election")
    
class ExtendElection(View):
    @method_decorator(login_required)
    def get(self, request, pk):
        try:
            election = get_object_or_404(Election, pk=pk)
            if election.end_date:
                election.end_date = election.end_date.strftime("%Y-%m-%dT%H:%M")
        except Election.DoesNotExist as e:
            logger.error(f"election does not exist : {e}")
            messages.error(request, "Une erreur inattendue s'est produite !")
            return redirect("election")

        return render(request, 'administration/extend_election.html', {
            'election': election
        })

    def post(self, request, pk):
        end_date = timezone.datetime.fromisoformat(request.POST.get("end_date"))
        local_tz = pytz.timezone('Africa/Kinshasa')
        local_end_date = local_tz.localize(end_date)
        end_date_utc = local_end_date.astimezone(pytz.utc)
        
        try:
            election = get_object_or_404(Election, pk=pk)

            if end_date_utc <= election.end_date:
                messages.error(request, "La date de fin doit être postérieure à la date fin avant !")
                return redirect("extend_election", pk=pk)
            else:
                election.end_date = local_end_date
                election.save()
                logger.info(f"L'élection {election} a été prolongée avec succès")
                messages.success(request, "L'élection a été prolongée avec succès !")
                return redirect("election")
        except Election.DoesNotExist as e:
            logger.error(f"L'élection n'existe pas : {e}")
            messages.error(request, "Une erreur inattendue s'est produite !")
            return redirect("election")

@login_required
def closeVoting(request, pk):
    try:
        election = get_object_or_404(Election, pk=pk)
        election.status = "closed"
        election.save()
        logger.info(f"L'élection {election} a été clôturée avec succès.")
        messages.success(request, "L'élection a été clôturée avec succès !")
        return redirect("election")
    except Election.DoesNotExist as e:
        logger.error(f"L'élection n'existe pas : {e}")
        messages.error(request, "Une erreur inattendue s'est produite !")
        return redirect("election")

@login_required
def publishedVoting(request, pk):
    try:
        with transaction.atomic():
            election = get_object_or_404(Election, pk=pk)
            votes = Vote.objects.filter(election=election)
            logger.info(f"Démarrage de la validation des votes pour {election}...")
            for vote in votes:
                logger.info(f"Validation du vote {vote}")
                elector_token = vote.elector_token_decrypt
                
                if not Elector.objects.filter(elector_token=elector_token).exists():
                    logger.warning(f"Le vote {vote} a un jeton invalide. Ce vote n'est pas authentique.")
                    vote.is_counted = True
                    vote.save()
                    continue
                if vote.is_blank:
                    logger.warning(f"Le vote {vote} est un vote blanc.")
                    vote.is_counted = True
                    vote.save()
                    continue
                if not (election.start_date <= vote.timestamp <= election.end_date):
                    logger.warning(f"Le vote {vote} est en dehors de la période de l'élection.")
                    vote.is_counted = True
                    vote.save()
                    continue
                if Vote.objects.filter(Q(election=election) & Q(elector_token=vote.elector_token)).count() > 1:
                    logger.warning(f"Le vote {vote} est dupliqué.")
                    vote.is_counted = True
                    vote.save()
                    continue
                
                vote.is_valid = True
                vote.is_counted = True
                vote.save()
                logger.info(f"Vote {vote} validé avec succès.")
            
            election.status = "published"
            election.save()
            logger.info(f"L'élection {election} a été publiée avec succès.")
            return redirect("election")
                
    except Election.DoesNotExist as e:
        logger.error(f"L'élection n'existe pas : {e}")
        messages.error(request, "Une erreur inattendue s'est produite !")
        return redirect("election")

@login_required
def detail_election(request, pk):
    try:
        election = get_object_or_404(Election, pk=pk)
        candidates = Candidate.objects.filter(election=election)
        candidates_exist = candidates.exists()
        if election.status == "published":
            total_votes = Vote.objects.filter(election=election, is_valid=True, is_counted=True).count()
            total_invalid_votes = Vote.objects.filter(election=election, is_valid=False, is_counted=True).count()
            total_blank_votes = Vote.objects.filter(election=election, is_blank=True, is_counted=True).count()
            
            count_electoral_list = ElectoralList.objects.all().count()
            electors_count = election.electors_count
            global_votes = Vote.objects.filter(election=election).count()
            participation_rate = (global_votes * 100) / count_electoral_list if count_electoral_list > 0 else 0

            resultat_pourcentage = calculer_resultats(request, election, Vote, Candidate)
            votes_pays = decompte_votes_par_pays(election, Vote, Candidate)

        else:
            total_votes = 0
            total_blank_votes = 0
            participation_rate = 0
            total_invalid_votes = 0
            resultat_pourcentage = []
            votes_pays = []
    except Election.DoesNotExist as e:
        logger.error(f"L'élection n'existe pas : {e}")
        messages.error(request, "Une erreur inattendue s'est produite !")
        return redirect("election")

    return render(request, "administration/consult_election.html", {
        "name": "election",
        "election": election,
        "candidates_exist": candidates_exist,
        "candidates": candidates,
        'total_votes': total_votes,
        'participation_rate': participation_rate,
        'total_blank_votes': total_blank_votes,
        'total_invalid_votes': total_invalid_votes,
        'resultat_pourcentage': resultat_pourcentage,
        'votes_pays': votes_pays
    })